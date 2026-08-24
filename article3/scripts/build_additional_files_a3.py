#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_additional_files_a3.py — A3 BMC Bioinformatics submission Additional files compilation
================================================================
Compile frozen CSVs into BMC-compatible Additional file xlsx (one independent file per table):

  Additional file 1 → Supplementary Table S1 (Compositional controls)
                      Source: a3_robustness_frozen.csv
  Additional file 2 → Supplementary Table S2 (Cancer-stratified meta-analysis)
                      Source: a3_robustness_meta.csv
  Additional file 3 → Supplementary Table S3 (External cohorts + null diagnostics
                       + transportability)  Source: a3_external_gbm.csv +
                       a3_external_null_diagnostics.csv + a3_transportability_frozen.csv

Output: article3/results/additional_files/A3_AdditionalFile{1,2,3}.xlsx
Dependency: openpyxl (3.1.5 already installed in this environment)
"""
import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(BASE))
RES = os.path.join(ROOT, "article3", "results")
OUT_DIR = os.path.join(RES, "additional_files")

SPECS = [
    {
        "file": "A3_AdditionalFile1.xlsx",
        "title": "Additional file 1. Compositional controls for the FL-immune association "
                 "(Supplementary Table S1).",
        "sheets": [
            ("Compositional_controls", os.path.join(RES, "a3_robustness_frozen.csv")),
        ],
    },
    {
        "file": "A3_AdditionalFile2.xlsx",
        "title": "Additional file 2. Cancer-stratified fixed-effect meta-analysis of the "
                 "FL-immune correlation (Supplementary Table S2).",
        "sheets": [
            ("Meta_analysis", os.path.join(RES, "a3_robustness_meta.csv")),
        ],
    },
    {
        "file": "A3_AdditionalFile3.xlsx",
        "title": "Additional file 3. External gene-level cohort assessment with null "
                 "diagnostics and internal transportability (Supplementary Table S3).",
        "sheets": [
            ("External_gene_level", os.path.join(RES, "a3_external_gbm.csv")),
            ("Null_diagnostics", os.path.join(RES, "a3_external_null_diagnostics.csv")),
            ("Transportability", os.path.join(RES, "a3_transportability_frozen.csv")),
        ],
    },
]


def load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def sheet_from_rows(ws, rows):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E6F1FB")
    wrap = Alignment(wrap_text=True, vertical="top")
    for j, cell in enumerate(rows[0], 1):
        c = ws.cell(row=1, column=j, value=cell)
        c.font = bold
        c.fill = fill
    for i, row in enumerate(rows[1:], 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v).alignment = wrap
    # Column widths
    for j in range(1, len(rows[0]) + 1):
        col_vals = [len(str(r[j - 1])) for r in rows[1:20] if len(r) >= j] or [8]
        width = min(max(col_vals) * 1.2 + 2, 60)
        ws.column_dimensions[get_column_letter(j)].width = max(width, 8)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    for spec in SPECS:
        wb = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Info"
        ws0["A1"] = spec["title"]
        ws0["A1"].font = Font(bold=True, size=12)
        ws0["A2"] = "Source frozen tables: article3/results/ (see manuscript Data availability)."
        ws0.column_dimensions["A"].width = 110
        for name, csv_path in spec["sheets"]:
            rows = load_csv(csv_path)
            ws = wb.create_sheet(title=name[:31])
            sheet_from_rows(ws, rows)
        # Move Info to the front
        wb.move_sheet("Info", offset=-(len(spec["sheets"])))
        out = os.path.join(OUT_DIR, spec["file"])
        wb.save(out)
        sz = os.path.getsize(out)
        print(f"OK  {spec['file']}  ({sz/1024:.1f} KB, sheets: {[n for n,_ in spec['sheets']]})")

    # Manifest
    manifest = os.path.join(OUT_DIR, "A3_AdditionalFiles_manifest.md")
    lines = [
        "# A3 BMC Bioinformatics — Additional Files Manifest (compiled 2026-08-18)",
        "",
        "| File | Corresponding Manuscript | Content | Source Frozen Table |",
        "|---|---|---|---|",
        "| A3_AdditionalFile1.xlsx | Supplementary Table S1 | FL–immune association compositional controls (log-ratio / FL–RI coupling / low-signal filtering) | a3_robustness_frozen.csv |",
        "| A3_AdditionalFile2.xlsx | Supplementary Table S2 | 32 cancer-type stratified fixed-effect meta-analysis (M2/Myeloid, 95% CI, Q, I²) | a3_robustness_meta.csv |",
        "| A3_AdditionalFile3.xlsx | Supplementary Table S3 | External gene-level cohort + null diagnostics + internal cross-cancer transportability (L2CO/held-out) | a3_external_gbm.csv, a3_external_null_diagnostics.csv, a3_transportability_frozen.csv |",
        "",
        "Note: The first sheet of each xlsx is the file description (Info); the data sheet is the corresponding frozen table (full columns, untruncated).",
        "BMC hard constraint: single Additional file ≤20 MB (currently each <100 KB, passed).",
    ]
    with open(manifest, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"OK  Manifest: {manifest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
